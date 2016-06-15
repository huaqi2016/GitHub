#!/usr/bin/env_python
# -*- coding: utf-8 -*-
#coding=utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
__author__ = 'zhoushengqiang'


class excel_process(object):
    def __init__(self):
        pass

    #对第colnum列进行写操作此模块读取和写入excel时，A1格的角标是[1,1]
    def readExcel(self, path):
        lineList=[]   #读取excel，将所有的数据放到一个list中
        dataList=[]   #按行为单位，将一行作为list一个元素
        paramList=[]  #以字典格式存储
        wb = load_workbook(path)
        #显示有多少张表
        #print "Worksheet range(s):", wb.get_named_ranges()
        #print "Worksheet name(s):", wb.get_sheet_names()
        #取第一张表
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])
        #显示表名，表行数，表列数
        #print "Work Sheet Titile:",ws.title
        row = ws.max_row
        col = ws.max_column
        for i in range(0,row):
            for j in range(0,col):
                if ws.rows[i][0] is None:
                    break
                lineList.append(ws.rows[i][j].value)
        for i in range(0,len(lineList),col):
            temp = lineList[i:i+col]
            dataList.append(temp)
        for i in range(1, len(dataList)):
            dictParam = dict(zip(*[dataList[0], dataList[i]]))
            paramList.append(dictParam)
        print(__file__,paramList)
        return paramList

    #对第colnum列进行写操作，此模块在读取和写入excel时，A1格的角标是[1,1]
    def writeColExcel(self, path, colnum, values):
        ids=[]
        wb = load_workbook(path)
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])
        ids = [i for i in range(2, len(values)+2)]
        data= dict(zip(ids, values))
        print data
        for i in range(2,len(values)+2):
            ws.cell(row=i, column=colnum).value = data[i]
        wb.save(path)

    #对第colnum列进行写操作此模块读取和写入excel时，A1格的角标是[1,1]
    def readExcelExp(self, path, tabname):
        lineList=[]   #读取excel，将所有的数据放到一个list中
        dataList=[]   #按行为单位，将一行作为list一个元素
        paramList=[]  #以字典格式存储
        expectList=[] #期望列表
        expmodList=[]
        wb = load_workbook(path)
        #显示有多少张表
        #print "Worksheet range(s):", wb.get_named_ranges()
        #print "Worksheet name(s):", wb.get_sheet_names()
        #取第一张表
        sheetnames = wb.get_sheet_names()
        #print sheetnames
        ws = wb.get_sheet_by_name(tabname)
        #显示表名，表行数，表列数
        #print "Work Sheet Titile:",ws.title
        row = ws.max_row
        col = ws.max_column
        for i in range(0,row):
            for j in range(0,col):
                if ws.rows[i][0].value is None:
                    break
                if ws.rows[i][j].value is None:
                    ws.rows[i][j].value=''
                lineList.append(ws.rows[i][j].value)

        for i in range(0,len(lineList),col):
            temp = lineList[i:i+col]
            dataList.append(temp)
        for i in range(1, len(dataList)):
            dictParam = dict(zip(*[dataList[0], dataList[i]]))
            expectList.append(dictParam['expect'])
            dictParam.pop('expect')
            #print dictParam
            paramList.append(dictParam)
        #print expectList
        for i in range(10, len(expectList)+10):
            expmodList.append(str(i)+str(expectList[i-10]))
        ret = dict(zip(expmodList, paramList))
        #print ret
        return ret

    #对第colnum列进行写操作，此模块在读取和写入excel时，A1格的角标是[1,1]
    def writeColExcelExp(self, path, colnum, values):
        ids=[]
        wb = load_workbook(path)
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])
        ids = [i for i in range(2, len(values)+2)]
        data= dict(zip(ids, values))
        print data
        for i in range(2,len(values)+2):
            ws.cell(row=i, column=colnum).value = data[i]
        wb.save(path)


if __name__ == '__main__':
    excelList=[]
    ep = excel_process()
    excelList=ep.readExcelExp('E:\\script\\test\\robotframework\case\\album.xlsx', 'AlbumCreate')
    #for i in range(0, len(excelList), 3):
    #    print excelList[i:i+3]
    #v=[165, 166, 167, 168, 169]
    #ep.writeColExcel('E:\\script\\test\\oauth\\case\\group_create.xlsx', 1, v)
